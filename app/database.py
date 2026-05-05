import sqlite3
from datetime import datetime, timedelta


def _connect(db_path):
    conn = sqlite3.connect(db_path, timeout=30)
    try:
        conn.execute('PRAGMA busy_timeout=30000')
        conn.execute('PRAGMA journal_mode=WAL')
    except sqlite3.OperationalError:
        # Si la DB está bloqueada momentáneamente, continuamos con la conexión.
        pass
    return conn


def init_db(db_path, retention_days=30):
    """Crea la tabla de logs si no existe e implementa rotación."""
    with _connect(db_path) as conn:
        try:
            conn.execute('''
                CREATE TABLE IF NOT EXISTS logs (
                    id        INTEGER PRIMARY KEY AUTOINCREMENT,
                    fecha     TEXT NOT NULL,
                    hora      TEXT NOT NULL,
                    modulo    TEXT NOT NULL,
                    archivo   TEXT NOT NULL,
                    tamano_kb REAL NOT NULL,
                    resultado TEXT NOT NULL,
                    detalle   TEXT
                )
            ''')
            # Crear índice en fecha para búsquedas rápidas
            try:
                conn.execute('CREATE INDEX IF NOT EXISTS idx_logs_fecha ON logs(fecha)')
            except sqlite3.OperationalError:
                pass
            conn.commit()
        except sqlite3.OperationalError as e:
            if 'locked' not in str(e).lower():
                raise
    
    # Limpiar logs antiguos
    limpiar_logs_antiguos(db_path, retention_days)


def limpiar_logs_antiguos(db_path, retention_days=30):
    """Elimina registros de logs más antiguos que retention_days."""
    fecha_limite = (datetime.now() - timedelta(days=retention_days)).strftime('%Y-%m-%d')
    with _connect(db_path) as conn:
        try:
            cursor = conn.execute(
                'DELETE FROM logs WHERE fecha < ?',
                (fecha_limite,)
            )
            eliminados = cursor.rowcount
            if eliminados > 0:
                conn.commit()
                print(f'[db] {eliminados} registro(s) de log eliminado(s) (> {retention_days} días)')
        except sqlite3.OperationalError as e:
            if 'locked' not in str(e).lower():
                raise


def registrar_log(db_path, modulo, archivo, tamano_bytes, resultado, detalle=None):
    """Inserta un registro en la tabla de logs."""
    now = datetime.now()
    with _connect(db_path) as conn:
        conn.execute(
            '''INSERT INTO logs (fecha, hora, modulo, archivo, tamano_kb, resultado, detalle)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (
                now.strftime('%Y-%m-%d'),
                now.strftime('%H:%M:%S'),
                modulo,
                archivo,
                round(tamano_bytes / 1024, 2),
                resultado,
                detalle,
            )
        )
        conn.commit()


def obtener_logs(db_path, limit=200):
    """Retorna los últimos N logs para consulta interna."""
    with _connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            'SELECT * FROM logs ORDER BY id DESC LIMIT ?', (limit,)
        ).fetchall()
    return [dict(r) for r in rows]


def exportar_logs_excel(db_path, ruta_salida=None):
    """
    Exporta todos los logs a un único archivo Excel con histórico completo.
    Una sola hoja 'Logs' con todos los registros ordenados por fecha/hora.
    Si ruta_salida es None, retorna bytes del Excel.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    with _connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            'SELECT * FROM logs ORDER BY fecha ASC, hora ASC'
        ).fetchall()
    rows = [dict(r) for r in rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Histórico Logs'

    # Cabecera
    headers = ['#', 'Fecha', 'Hora', 'Módulo', 'Archivo', 'Tamaño (KB)', 'Resultado', 'Detalle']
    header_fill = PatternFill(start_color='1B4FD8', end_color='1B4FD8', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 20

    # Filas de datos
    fill_ok  = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
    fill_err = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    font_ok  = Font(color='16A34A', bold=True, size=9)
    font_err = Font(color='DC2626', bold=True, size=9)

    for row_idx, r in enumerate(rows, start=2):
        es_ok = r['resultado'] == 'exito'
        fill  = fill_ok if es_ok else fill_err
        valores = [
            r['id'], r['fecha'], r['hora'], r['modulo'],
            r['archivo'], r['tamano_kb'], r['resultado'], r['detalle'] or ''
        ]
        for col, val in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            if col == 7:  # columna resultado
                cell.font = font_ok if es_ok else font_err

    # Anchos de columna
    anchos = [6, 12, 10, 14, 35, 13, 12, 30]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    # Filtro automático y fila fija
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:H{len(rows) + 1}'

    if ruta_salida:
        wb.save(ruta_salida)
        return None
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
